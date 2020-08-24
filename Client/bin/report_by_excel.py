# -*- coding:utf-8 -*-
import json
import urllib.request
import urllib.parse
import os
import sys
import time
import openpyxl
import xlrd

def update_test(data):
    """
    创建测试用例
    :return:
    """
    # 将数据打包到一个字典内，并转换为json格式
    datas = {"asset_data": json.dumps(data)}
    # 根据settings中的配置，构造url
    url = "http://%s:%s%s" % (Params['server'], Params['port'], Params['url'])
    print('正在将数据发送至： [%s]  ......' % url)
    try:
        # 使用Python内置的urllib.request库，发送post请求。
        # 需要先将数据进行封装，并转换成bytes类型
        data_encode = urllib.parse.urlencode(datas).encode()
        response = urllib.request.urlopen(url=url, data=data_encode, timeout=Params['request_timeout'])
        print("\033[31;1m发送成功+1！\033[0m ")
        message = response.read().decode()
        print("返回结果：%s" % message)
    except Exception as e:
        message = "发送失败"
        print("\033[31;1m发送失败，%s\033[0m" % e)


def pack_server(filename):
    wb = xlrd.open_workbook(filename)
    server = wb.sheet_by_name('server')
    ser_nor = server.nrows
    ser_nol = server.ncols
    ser_dict = {}
    for i in range(1, ser_nor):
        for j in range(ser_nol):
            title = server.cell_value(0, j)
            value = server.cell_value(i, j)
            ser_dict[title] = value
        yield ser_dict

def pack_security(filename):
    wb = xlrd.open_workbook(filename)
    security = wb.sheet_by_name('security')
    sec_nor = security.nrows
    sec_nol = security.ncols
    sec_dict = {}
    for i in range(1, sec_nor):
        for j in range(sec_nol):
            title = security.cell_value(0, j)
            value = security.cell_value(i, j)
            sec_dict[title] = value
        yield sec_dict


def pack_test(file, data):
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_sheet_by_name('server')
    ru = sheet.max_row  # 获取行数
    cu = sheet.max_column  # 获取列数
    for row in sheet.iter_rows():
        if row[1].value == "server":
            # data = {}
            runtime = time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
            data["aud_sn"] = runtime
            x = row[1].coordinate
            y = row[1].value
            print(x)
            print(y)


if __name__ == '__main__':
    filename = sys.argv[1]
    Params = {
        "server": "127.0.0.1",
        "port": 8000,
        'url': '/assets/report/',
        'request_timeout': 30,
    }
    # data = {
    #         "aud_sn": "Ser-025",    # 资产序列号（唯一）
    #         "aud_asset_type": "server",    # 资产类型（server-服务器，securitydevice-安全设备，networkdevice-网络设备，otherdevice-其它设备）
    #         "aud_name": "WEB通行证",    # 资产名称
    #         "aud_manage_ip": "10.0.0.1",    # 管理IP
    #         "aud_in_ip": "192.168.0.1",    # 内网IP
    #         "aud_out_ip": "202.106.0.1",    # 外网IP
    #         "aud_visit_url": "www.xxx.com",    # 域名/URL
    #         "aud_url_status": "200",    # 状态码
    #         "aud_business_style": "web",    # 业务类型
    #         "aud_department_per": "鹿晗",    # 总负责人
    #         "aud_phone": "13829328231",    # 总负责人-电话
    #         "aud_mail": "luhan@test.com",    # 总负责人-邮箱
    #         "aud_regional": "DMZ区",    # 部署区域
    #         "aud_deployment": "串接",    # 部署方式
    #         "aud_status": "0",    # 资产状态（0-在线，1-下线，2-未知，3-故障，4-备用)
    #         "aud_le_status": "5",    # 重要程度（5-非常高，4-高，3-中，2-低，1-可忽略）
    #         "aud_memo": "上线时间-2020年1月1日",    # 备注
    #         "ser_sub_asset_type": "1",    # 服务器类型（0-物理服务器，1-虚拟机，2-其它设备）
    #         "ser_os_type": "CentOS",    # 操作系统类型
    #         "ser_os_release": "7.1",    # 操作系统版本
    #         "ser_model": "Core3.2312.41",    # 内核信息
    #         "ser_department_per": "梁山伯",    # 系统负责人
    #         "ser_phone": "13287463134",    # 系统负责人-电话
    #         "ser_mail": "2sdfoe@test.com",    # 系统负责人-邮箱
    #         "ser_memo": "需要重新安装操作系统",    # 服务器备注
    #         "idc_name": "苏州桥机房",    # 机房名称
    #         "idc_link": "移动",    # 链路信息
    #         "idc_cab": "第1机柜",    # 机柜名称
    #         "idc_num": "bcs-123",    # 编号信息
    #         "idc_memo": "新购02",    # 机房备注
    #         "serv_ser": "web|tomcat|mysql|",    # 服务
    #         "serv_port": "80|8080|3306",    # 开放端口
    #         "serv_visit": "www.csdfsfasv.com|Null",    # 访问方式
    #         "serv_memo": "需要更换熟知端口",    # 服务备注
    #         "data_model": "MongDB",    # 数据库类型
    #         "data_version": "ss32.123",    # 数据库版本
    #         "data_department_per": "宋茜",    # 数据库负责人
    #         "data_phone": "18721341241",    # 数据库负责人-电话
    #         "data_mail": "songqian@test.com",    # 数据库负责人-邮箱
    #         "data_memo": "需优化",    # 数据库备注
    #         "mid_model": "TeTomcat",    # 中间件类型
    #         "mid_version": "8.3.1",    # 中间件版本
    #         "mid_department_per": "赵四",    # 中间件负责人
    #         "mid_phone": "17723123122",    # 中间件负责人-电话
    #         "mid_mail": "chzhoaben@test.com",    # 中间件负责人-邮箱
    #         "mid_memo": "计划下线时间2月",    # 中间件备注
    #         "man_name": "奇安信",    # 服务商
    #         "man_frame": "Django",    # 框架
    #         "man_language": "Java",    # 开发语言
    #         "man_version": "23.123",    # 版本
    #         "man_department": "系统开发部",    # 服务商部门
    #         "man_department_per": "杨开慧",    # 服务商负责人
    #         "man_phone": "12831412314",    # 服务商负责人-电话
    #         "man_mail": "2o3423@test.com",    # 服务商负责人-邮箱
    #         "man_memo": "安全新高峰"    # 服务商备注
    #         }
    for ser in pack_server(filename):
        runtime = time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
        ser["aud_sn"] = runtime
        update_test(ser)
        time.sleep(1)
    for sec in pack_security(filename):
        runtime = time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
        sec["aud_sn"] = runtime
        update_test(sec)
        time.sleep(1)
    # pack_test(file, data)



